"""Moulinette config loader — extracts all configuration from PMS_Loader.xlsm.

DEPRECATED: This module is superseded by app.services.importer which
writes configuration directly to SQLite. Kept for reference only.
Use `importer.import_moulinette()` instead of `MoulinetteLoader.load()`.
"""

import warnings as _warnings
_warnings.warn(
    "moulinette_loader is deprecated — use app.services.importer instead",
    DeprecationWarning,
    stacklevel=2,
)

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from app.models.schemas import (
    ColumnSchema,
    ConcessionConfig,
    DPRNamingRule,
    MappingEntry,
    MoulinetteConfig,
    QCRule,
    UOMEntry,
)
from app.services.logger import LogService


# Singleton config store
_config: MoulinetteConfig | None = None


def get_moulinette_config() -> MoulinetteConfig | None:
    return _config


class MoulinetteLoader:
    """Reads PMS_Loader .xlsm and extracts all config tables."""

    def __init__(self) -> None:
        self.logger = LogService.get()

    async def load(self, path: Path) -> MoulinetteConfig:
        global _config

        await self.logger.info(
            f"Loading moulinette config from {path.name}", source="moulinette"
        )

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

        try:
            config = MoulinetteConfig(
                schema_dc=self._load_schema(wb, "DatafileDD", "B", "A"),
                schema_dw=self._load_schema(wb, "DatafileDD", "J", "I"),
                schema_mc=self._load_schema(wb, "DatafileDD", "F", "E"),
                schema_wt=self._load_schema(wb, "DatafileDD", "N", "M"),
                uom_entries=self._load_uom(wb),
                concessions=self._load_concessions(wb),
                naming_rules=self._load_naming_rules(wb),
                qc_rules=self._load_qc_rules(wb),
                template_mappings=self._load_template(wb),
            )

            # Load parameters
            params = self._load_parameters(wb)
            config.mapping_path = params.get("mapping_path", "")
            config.dpr_path = params.get("dpr_path", "")
            config.output_path = params.get("output_path", "")
            config.sm3_to_nm3 = params.get("sm3_to_nm3", 0.947916)
            config.nm3_to_sm3 = params.get("nm3_to_sm3", 1.05494579688496)

        finally:
            wb.close()

        _config = config

        await self.logger.success(
            f"Moulinette loaded: {len(config.schema_dc)} DC cols, "
            f"{len(config.schema_dw)} DW cols, "
            f"{len(config.concessions)} concessions, "
            f"{len(config.uom_entries)} UOM entries, "
            f"{len(config.qc_rules)} QC rules, "
            f"{len(config.template_mappings)} template mappings",
            source="moulinette",
        )

        return config

    # ── DatafileDD: Column schemas ─────────────────────────────────────

    def _load_schema(
        self, wb: Any, sheet: str, code_col: str, attr_col: str
    ) -> list[ColumnSchema]:
        """Load a column schema from DatafileDD.

        Each schema is stored as two columns: attribute name and code.
        The column index is derived from the row order.
        """
        if sheet not in wb.sheetnames:
            return []

        ws = wb[sheet]
        schemas: list[ColumnSchema] = []
        code_idx = openpyxl.utils.column_index_from_string(code_col)
        attr_idx = openpyxl.utils.column_index_from_string(attr_col)

        for row in ws.iter_rows(min_row=2, max_col=max(code_idx, attr_idx) + 1):
            code_cell = row[code_idx - 1] if code_idx - 1 < len(row) else None
            attr_cell = row[attr_idx - 1] if attr_idx - 1 < len(row) else None

            if code_cell and code_cell.value and attr_cell and attr_cell.value:
                code = str(code_cell.value).strip()
                attr = str(attr_cell.value).strip()
                if code and attr:
                    schemas.append(ColumnSchema(
                        attribute=attr,
                        code=code,
                        column_index=len(schemas) + 1,
                    ))
            elif not code_cell or not code_cell.value:
                # End of this schema block
                if schemas:
                    break

        return schemas

    # ── UOM conversion table ───────────────────────────────────────────

    def _load_uom(self, wb: Any) -> list[UOMEntry]:
        if "UOM" not in wb.sheetnames:
            return []

        ws = wb["UOM"]
        entries: list[UOMEntry] = []

        for row in ws.iter_rows(min_row=2, max_col=3):
            unit = row[0].value
            factor = row[1].value
            target = row[2].value if len(row) > 2 else ""

            if unit and factor is not None:
                try:
                    entries.append(UOMEntry(
                        unit=str(unit).strip().upper(),
                        factor=float(factor),
                        target_unit=str(target or "").strip(),
                    ))
                except (ValueError, TypeError):
                    continue

        return entries

    # ── Concession list from Parameters D1 ─────────────────────────────

    def _load_concessions(self, wb: Any) -> list[ConcessionConfig]:
        if "Parameters" not in wb.sheetnames:
            return []

        ws = wb["Parameters"]
        concessions: list[ConcessionConfig] = []

        # D column = 4, parameters start at row 2
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=9):
            name = row[0].value  # col D
            if not name or str(name).strip() == "":
                continue

            active_d = row[1].value if len(row) > 1 else 0  # col E
            active_m = row[2].value if len(row) > 2 else 0  # col F
            active_wt = row[3].value if len(row) > 3 else 0  # col G
            monthly_rpt = row[4].value if len(row) > 4 else ""  # col H
            mapping = row[5].value if len(row) > 5 else ""  # col I

            concessions.append(ConcessionConfig(
                name=str(name).strip(),
                active_daily=str(active_d) == "1",
                active_monthly=str(active_m) == "1",
                active_well_test=str(active_wt) == "1",
                monthly_report=str(monthly_rpt or "").strip(),
                mapping_file=str(mapping or "").strip(),
            ))

        return concessions

    # ── DPR naming rules from Parameters K1 ────────────────────────────

    def _load_naming_rules(self, wb: Any) -> list[DPRNamingRule]:
        if "Parameters" not in wb.sheetnames:
            return []

        ws = wb["Parameters"]
        rules: list[DPRNamingRule] = []

        # K=11, L=12, M=13, N=14, O=15, P=16, Q=17, R=18
        for row in ws.iter_rows(min_row=2, min_col=11, max_col=18):
            alias = row[0].value  # col K
            if not alias or str(alias).strip() == "":
                continue

            file_alias = row[1].value if len(row) > 1 else alias  # col L
            ext = row[2].value if len(row) > 2 else "xlsx"  # col M
            fmt = row[3].value if len(row) > 3 else "ddmmyyyy"  # col N
            left = row[4].value if len(row) > 4 else ""  # col O
            right = row[5].value if len(row) > 5 else ""  # col P
            suffix = row[6].value if len(row) > 6 else ""  # col Q
            prefix = row[7].value if len(row) > 7 else ""  # col R

            rules.append(DPRNamingRule(
                alias=str(alias).strip(),
                file_alias=str(file_alias or alias).strip(),
                extension=str(ext or "xlsx").strip(),
                date_format=str(fmt or "ddmmyyyy").strip(),
                left_sep=str(left or "").strip(),
                right_sep=str(right or "").strip(),
                suffix=str(suffix or "").strip(),
                prefix=str(prefix or "").strip(),
            ))

        return rules

    # ── QC cleaning rules ──────────────────────────────────────────────

    def _load_qc_rules(self, wb: Any) -> list[QCRule]:
        if "QC_CLeaning" not in wb.sheetnames:
            return []

        ws = wb["QC_CLeaning"]
        rules: list[QCRule] = []

        for row in ws.iter_rows(min_row=2, max_col=5):
            sheet = row[0].value
            col_range = row[1].value
            search = row[2].value
            replace = row[3].value if len(row) > 3 else ""
            active = row[4].value if len(row) > 4 else None

            if sheet and col_range is not None:
                rules.append(QCRule(
                    sheet=str(sheet).strip(),
                    column_range=str(col_range).strip(),
                    search_value=str(search) if search is not None else "",
                    replace_value=str(replace) if replace is not None else "",
                    active=active != 0 if active is not None else True,
                ))

        return rules

    # ── Template mappings ──────────────────────────────────────────────

    def _load_template(self, wb: Any) -> list[MappingEntry]:
        if "Template" not in wb.sheetnames:
            return []

        ws = wb["Template"]
        entries: list[MappingEntry] = []

        # B=2, C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10, K=11
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=11):
            file_alias = row[0].value   # B - file name
            sheet_name = row[1].value   # C - sheet name
            well_name = row[2].value    # D - well name operator
            ubhi = row[3].value         # E - UBHI
            completion = row[4].value   # F - completion name
            template = row[5].value     # G - template type (DC/DW/MC/WT)
            attribute = row[6].value    # H - attribute name
            attr_code = row[7].value    # I - attribute code
            cell_ref = row[8].value     # J - mapping ref
            unit = row[9].value         # K - unit

            if not file_alias or not template:
                continue

            entries.append(MappingEntry(
                file_alias=str(file_alias).strip(),
                sheet_name=str(sheet_name or "").strip(),
                well_name=str(well_name or "").strip(),
                ubhi=str(ubhi or "").strip(),
                completion=str(completion or "").strip(),
                template_type=str(template).strip().upper(),
                attribute=str(attribute or "").strip(),
                attribute_code=str(attr_code or "").strip(),
                cell_ref=str(cell_ref or "").strip() if cell_ref and str(cell_ref) != "0" else "",
                unit=str(unit or "").strip() if unit and str(unit) != "0" else "",
            ))

        return entries

    # ── Scalar parameters from B column ────────────────────────────────

    def _load_parameters(self, wb: Any) -> dict[str, Any]:
        if "Parameters" not in wb.sheetnames:
            return {}

        ws = wb["Parameters"]
        params: dict[str, Any] = {}

        # Read B2..B9
        rows_map = {
            2: "mapping_path",
            3: "dpr_path",
            4: "output_path",
            5: "date",
            6: "source_path",
            7: "dest_path",
            8: "sm3_to_nm3",
            9: "nm3_to_sm3",
        }

        for row_num, key in rows_map.items():
            cell = ws.cell(row=row_num, column=2)
            val = cell.value
            if val is not None:
                if key in ("sm3_to_nm3", "nm3_to_sm3"):
                    try:
                        params[key] = float(val)
                    except (ValueError, TypeError):
                        pass
                else:
                    params[key] = str(val).strip()

        return params
