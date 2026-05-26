"""Cell reference reader — reads values from closed Excel workbooks.

Supports 3 reference types from the VBA moulinette:
1. Simple: "B15" — direct cell read
2. Calculated: "?B15+B16-B17" — formula with +, -, # operators
3. Multi-file: "![ALIAS/sheet]ref(factor)+..." — cross-file references

NOTE: Full implementation requires actual DPR files and mapping files.
This module provides the interface and parsing logic; the file-reading
methods will be completed when sample DPR files are available.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

from app.services.logger import LogService


# Cache of opened workbooks to avoid re-opening
_wb_cache: dict[str, Any] = {}


async def clear_cache() -> None:
    """Close all cached workbooks."""
    for wb in _wb_cache.values():
        try:
            wb.close()
        except Exception:
            pass
    _wb_cache.clear()


def _get_workbook(filepath: str) -> Any:
    """Get a workbook from cache or open it."""
    if filepath not in _wb_cache:
        path = Path(filepath)
        if not path.exists():
            return None
        _wb_cache[filepath] = openpyxl.load_workbook(
            str(path), read_only=True, data_only=True
        )
    return _wb_cache[filepath]


def read_cell(
    dpr_path: str, dpr_name: str, sheet_name: str, cell_ref: str
) -> Any:
    """Read a single cell value from a DPR Excel file.

    Equivalent to VBA GetValueFromDPRFile().
    """
    filepath = str(Path(dpr_path) / dpr_name)
    wb = _get_workbook(filepath)
    if wb is None:
        return ""

    try:
        # Find sheet by name, handling trailing whitespace in Excel
        ws = None
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Fuzzy match: strip both sides and compare case-insensitively
            target = sheet_name.strip().lower()
            for sn in wb.sheetnames:
                if sn.strip().lower() == target:
                    ws = wb[sn]
                    break
        if ws is None:
            return ""

        # Parse cell reference like "B15" into row/col
        match = re.match(r"([A-Za-z]+)(\d+)", cell_ref)
        if not match:
            return ""
        col = openpyxl.utils.column_index_from_string(match.group(1))
        row = int(match.group(2))
        val = ws.cell(row=row, column=col).value
        return val if val is not None else ""
    except Exception:
        return ""


def resolve_ref(
    ref: str,
    dpr_path: str,
    dpr_name: str,
    sheet_name: str,
    dpr_list: dict[str, str] | None = None,
) -> Any:
    """Resolve a cell reference of any type.

    Dispatches to the appropriate parser based on prefix:
    - "?" → calculated reference (sum/subtract/concat)
    - "!" → multi-file reference
    - plain → simple cell reference
    """
    ref = ref.strip()

    if not ref or ref == "0":
        return ""

    # Remove colon if present (VBA does Replace(ref, ":", ""))
    ref = ref.replace(":", "")

    if ref.startswith("?"):
        return _parse_calculated(ref, dpr_path, dpr_name, sheet_name)
    elif ref.startswith("!"):
        return _parse_multi_file(ref, dpr_path, dpr_list or {})
    else:
        return read_cell(dpr_path, dpr_name, sheet_name, ref)


def _parse_calculated(
    ref: str, dpr_path: str, dpr_name: str, sheet_name: str
) -> Any:
    """Parse calculated reference: ?B15+B16-B17

    Operators: + (add), - (subtract), # (string concat with space)
    """
    result: Any = 0
    current_ref = ""
    operator = ""

    for i in range(1, len(ref)):  # skip the '?' prefix
        ch = ref[i]

        if ch in ("+", "-", "#"):
            # Process accumulated reference
            if current_ref:
                val = read_cell(dpr_path, dpr_name, sheet_name, current_ref)
                result = _apply_operator(result, val, operator)
                current_ref = ""
            operator = ch
        else:
            current_ref += ch

    # Process last reference
    if current_ref:
        val = read_cell(dpr_path, dpr_name, sheet_name, current_ref)
        result = _apply_operator(result, val, operator)

    return result


def _parse_multi_file(
    ref: str, dpr_path: str, dpr_list: dict[str, str]
) -> Any:
    """Parse multi-file reference: ![ALIAS/sheet]ref(factor)+...

    Format: ![wb_alias/sheet_name]cell_ref(conversion_factor)
    Multiple terms can be joined with +
    """
    result: Any = 0
    # Pattern: [alias/sheet]ref(factor)
    pattern = r"\[([^/]+)/([^\]]+)\]([A-Za-z]+\d+)(?:\(([^)]+)\))?"
    matches = re.findall(pattern, ref)

    for wb_alias, ws_name, cell_ref, factor_str in matches:
        wb_alias = wb_alias.strip()
        ws_name = ws_name.strip()

        # Resolve actual filename from alias
        dpr_name = dpr_list.get(wb_alias, "")
        if not dpr_name:
            continue

        val = read_cell(dpr_path, dpr_name, ws_name, cell_ref)

        # Apply inline conversion factor
        if factor_str and val and _is_numeric(val):
            try:
                val = float(val) * float(factor_str)
            except (ValueError, TypeError):
                pass

        # Accumulate
        if _is_numeric_or_empty(val) and _is_numeric_or_empty(result):
            result = float(result or 0) + float(val or 0)
        elif val:
            result = val

    return result


def _apply_operator(accumulator: Any, value: Any, operator: str) -> Any:
    """Apply an operator between accumulator and value."""
    if operator == "" or operator == "+":
        if _is_numeric_or_empty(accumulator) and _is_numeric_or_empty(value):
            return float(accumulator or 0) + float(value or 0)
        return value
    elif operator == "-":
        if _is_numeric_or_empty(accumulator) and _is_numeric_or_empty(value):
            return float(accumulator or 0) - float(value or 0)
        return accumulator
    elif operator == "#":
        # String concatenation with space
        return f"{accumulator} {value}".strip()
    return value


def _is_numeric(val: Any) -> bool:
    """Check if a value is actually numeric (not None/empty).

    Use this for type-checking contexts where you need to know if a
    value genuinely contains a number.
    """
    if val is None or val == "":
        return False
    try:
        float(val)
        return True
    except (ValueError, TypeError):
        return False


def _is_numeric_or_empty(val: Any) -> bool:
    """Check if a value can participate in arithmetic (including None/empty as 0).

    Use this in accumulation contexts where missing values should be
    treated as zero.
    """
    if val is None or val == "":
        return True  # Treat empty as 0 in arithmetic
    try:
        float(val)
        return True
    except (ValueError, TypeError):
        return False
