"""Import configuration from external Excel files into SQLite.

Supports two import sources:
1. PMS_Loader moulinette (*.xlsm) — imports concessions, UOM, QC rules, naming rules
2. Mapping files (Mapping_*.xlsx) — imports cell-reference mappings for a concession
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from app.models.config_models import (
    ConcessionCreate,
    ImportResult,
    QCRuleCreate,
    UOMEntryCreate,
)
from app.services import config_store
from app.services.database import get_connection
from app.services.logger import LogService


# ═══════════════════════════════════════════════════════════════════════
# Moulinette Import
# ═══════════════════════════════════════════════════════════════════════

async def import_moulinette(path: Path) -> ImportResult:
    """Import configuration from a PMS_Loader .xlsm workbook.

    Imports: concessions, UOM entries, QC rules, naming rules, parameters.
    Does NOT import cell mappings — those come from mapping files.
    """
    logger = LogService.get()
    await logger.info(f"Importing moulinette from {path.name}", source="importer")

    wb = openpyxl.load_workbook(str(path), data_only=True)
    result = ImportResult(source=path.name)

    try:
        # ── Parameters (B column) ──────────────────────────────────
        if "Parameters" in wb.sheetnames:
            ws = wb["Parameters"]
            param_map = {
                2: "mapping_path",
                3: "dpr_path",
                4: "output_path",
                8: "sm3_to_nm3",
                9: "nm3_to_sm3",
            }
            for row_num, key in param_map.items():
                val = ws.cell(row=row_num, column=2).value
                if val is not None:
                    config_store.set_parameter(key, str(val).strip())

            # ── Concessions (D-I columns) ──────────────────────────
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=9):
                name = row[0].value  # D
                if not name or str(name).strip() == "":
                    continue

                active_d = row[1].value if len(row) > 1 else 0  # E
                active_m = row[2].value if len(row) > 2 else 0  # F
                active_wt = row[3].value if len(row) > 3 else 0  # G
                monthly_rpt = row[4].value if len(row) > 4 else ""  # H
                mapping_file = row[5].value if len(row) > 5 else ""  # I

                conc_data = ConcessionCreate(
                    name=str(name).strip(),
                    active_daily=str(active_d) == "1",
                    active_monthly=str(active_m) == "1",
                    active_well_test=str(active_wt) == "1",
                    monthly_report=str(monthly_rpt or "").strip(),
                )

                try:
                    config_store.create_concession(conc_data)
                    result.concessions_imported += 1
                except Exception as e:
                    result.warnings.append(f"Concession {name}: {e}")

            # ── Naming rules (K-R columns) ─────────────────────────
            _import_naming_rules(ws, result)

        # ── UOM entries ────────────────────────────────────────────
        if "UOM" in wb.sheetnames:
            ws = wb["UOM"]
            for row in ws.iter_rows(min_row=2, max_col=3):
                unit = row[0].value
                factor = row[1].value
                target = row[2].value if len(row) > 2 else ""
                if unit and factor is not None:
                    try:
                        config_store.add_uom(UOMEntryCreate(
                            unit=str(unit).strip().upper(),
                            factor=float(factor),
                            target_unit=str(target or "").strip(),
                        ))
                        result.uom_imported += 1
                    except Exception:
                        pass

        # ── QC rules ───────────────────────────────────────────────
        if "QC_CLeaning" in wb.sheetnames:
            ws = wb["QC_CLeaning"]
            has_sheet_col = ws.max_column >= 5  # v2.8 has sheet/column context

            for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
                if has_sheet_col:
                    search = row[2].value if len(row) > 2 else None  # C
                    replace = row[3].value if len(row) > 3 else ""  # D
                    active = row[4].value if len(row) > 4 else 1  # E
                else:
                    search = row[0].value  # A
                    replace = row[1].value if len(row) > 1 else ""  # B
                    active = row[2].value if len(row) > 2 else 1  # C

                if search is not None:
                    try:
                        config_store.add_qc_rule(QCRuleCreate(
                            search_value=str(search),
                            replace_value=str(replace or ""),
                            active=active != 0 if active is not None else True,
                        ))
                        result.qc_rules_imported += 1
                    except Exception:
                        pass

        # ── DPR alias + naming from temp sheet ─────────────────────
        if "temp" in wb.sheetnames:
            _import_dpr_aliases_from_temp(wb["temp"], result)

    finally:
        wb.close()

    await logger.success(
        f"Moulinette import: {result.concessions_imported} concessions, "
        f"{result.uom_imported} UOM, {result.qc_rules_imported} QC rules",
        source="importer",
    )

    return result


def _import_naming_rules(ws: Any, result: ImportResult) -> None:
    """Import naming rules from Parameters K-R columns."""
    conn = get_connection()
    try:
        for row in ws.iter_rows(min_row=2, min_col=11, max_col=18):
            alias = row[0].value  # K
            if not alias or str(alias).strip() == "":
                continue

            file_alias = row[1].value if len(row) > 1 else alias  # L
            ext = row[2].value if len(row) > 2 else "xlsx"  # M
            fmt = row[3].value if len(row) > 3 else "ddmmyyyy"  # N
            left = row[4].value if len(row) > 4 else ""  # O
            right = row[5].value if len(row) > 5 else ""  # P
            suf = row[6].value if len(row) > 6 else ""  # Q
            pref = row[7].value if len(row) > 7 else ""  # R

            # Clean VBA date format strings
            fmt_str = str(fmt or "ddmmyyyy").strip()
            # Remove Excel locale prefixes like [$-en-US]
            if fmt_str.startswith("["):
                fmt_str = fmt_str.split("]")[-1].split(";")[0]

            conn.execute("""
                INSERT INTO naming_rules (alias, file_alias, extension, date_format,
                    left_sep, right_sep, prefix, suffix)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                str(alias).strip(),
                str(file_alias or alias).strip(),
                str(ext or "xlsx").strip(),
                fmt_str,
                str(left or "").strip(),
                str(right or "").strip(),
                str(pref or "").strip(),
                str(suf or "").strip(),
            ))
            result.naming_rules_imported += 1

        conn.commit()
    finally:
        conn.close()


def _import_dpr_aliases_from_temp(ws: Any, result: ImportResult) -> None:
    """Update concession DPR aliases from the temp sheet (cols D-K)."""
    conn = get_connection()
    try:
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=11):
            conc_name = row[0].value  # D = concession name
            dpr_file = row[1].value  # E = DPR file name
            dpr_alias = row[6].value if len(row) > 6 else ""  # J = DPR Alias
            fmt = row[7].value if len(row) > 7 else ""  # K = Format

            if not conc_name:
                continue

            name = str(conc_name).strip()
            alias = str(dpr_alias or "").strip()
            file_name = str(dpr_file or "").strip()

            # Clean format
            fmt_str = str(fmt or "").strip()
            if fmt_str.startswith("["):
                fmt_str = fmt_str.split("]")[-1].split(";")[0]

            # Try to match by name and update
            # Extract sheet name from the DPR file (will be set when mapping is imported)
            if alias or file_name:
                conn.execute("""
                    UPDATE concessions
                    SET dpr_file_alias = CASE WHEN ? != '' THEN ? ELSE dpr_file_alias END,
                        date_format = CASE WHEN ? != '' THEN ? ELSE date_format END
                    WHERE name = ?
                """, (alias, alias, fmt_str, fmt_str, name))

        conn.commit()
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════════════
# Mapping File Import
# ═══════════════════════════════════════════════════════════════════════

async def import_mapping_file(path: Path, concession_id: str | None = None) -> ImportResult:
    """Import cell-reference mappings from a Mapping_*.xlsx file.

    If concession_id is provided, maps to that concession.
    Otherwise, auto-detects from the file content.
    """
    logger = LogService.get()
    await logger.info(f"Importing mapping from {path.name}", source="importer")

    wb = openpyxl.load_workbook(str(path), data_only=True)
    result = ImportResult(source=path.name)

    try:
        # Auto-detect concession from file content
        detected_conc = None
        detected_sheet = None
        for sname in ["Mapping_concession", "Mapping_Well", "Mapping_Well_Test"]:
            if sname in wb.sheetnames:
                ws = wb[sname]
                row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
                if row2 and row2[0][0]:
                    detected_conc = str(row2[0][0]).strip()
                    detected_sheet = str(row2[0][2]).strip() if row2[0][2] else ""
                    break

        if not concession_id and detected_conc:
            # Find matching concession in DB
            concs = config_store.list_concessions()
            for c in concs:
                if detected_conc.lower() in c.name.lower() or c.name.lower() in detected_conc.lower():
                    concession_id = c.id
                    break

        if not concession_id and detected_conc:
            # Create the concession
            new_conc = config_store.create_concession(ConcessionCreate(
                name=detected_conc.upper() + " NEW",
                dpr_file_alias=detected_conc,
                dpr_sheet=detected_sheet or "",
            ))
            concession_id = new_conc.id
            result.concessions_imported = 1

        if not concession_id:
            result.warnings.append("Could not determine concession for this mapping file")
            return result

        # Update DPR sheet if detected
        if detected_sheet:
            config_store.update_concession(concession_id, type("U", (), {
                "name": None, "dpr_file_alias": None,
                "dpr_sheet": detected_sheet,
                "active_daily": None, "active_monthly": None,
                "active_well_test": None, "date_format": None,
                "monthly_report": None,
            })())

        mappings: list[dict] = []

        # ── Mapping_concession (DC + MC) ───────────────────────────
        if "Mapping_concession" in wb.sheetnames:
            ws = wb["Mapping_concession"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                # Cols: concession, file_name, sheet_name, Template, attribut, attribut_code, mapping-ref, unite
                template_type = str(row[3] or "").strip().upper()
                attr = str(row[4] or "").strip()
                code = str(row[5] or "").strip()
                ref = str(row[6] or "").strip() if row[6] else ""
                unit = str(row[7] or "").strip() if len(row) > 7 and row[7] else ""

                if code:
                    mappings.append({
                        "template_type": template_type,
                        "well_name": "",
                        "ubhi": "",
                        "completion": "",
                        "attribute_code": code,
                        "attribute": attr,
                        "cell_ref": ref,
                        "unit": unit,
                    })

        # ── Mapping_Well (DW) ──────────────────────────────────────
        if "Mapping_Well" in wb.sheetnames:
            ws = wb["Mapping_Well"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                # Cols: Concession, File, Sheet, WellName, UBHI, Completion, Template, Attribut, Code, Ref, Unite
                well_name = str(row[3] or "").strip() if len(row) > 3 else ""
                ubhi = str(row[4] or "").strip() if len(row) > 4 else ""
                completion = str(row[5] or "").strip() if len(row) > 5 else ""
                template_type = str(row[6] or "").strip().upper() if len(row) > 6 else "DW"
                attr = str(row[7] or "").strip() if len(row) > 7 else ""
                code = str(row[8] or "").strip() if len(row) > 8 else ""
                ref = str(row[9] or "").strip() if len(row) > 9 and row[9] else ""
                unit = str(row[10] or "").strip() if len(row) > 10 and row[10] else ""

                if code:
                    mappings.append({
                        "template_type": template_type,
                        "well_name": well_name,
                        "ubhi": ubhi,
                        "completion": completion,
                        "attribute_code": code,
                        "attribute": attr,
                        "cell_ref": ref,
                        "unit": unit,
                    })

        # ── Mapping_Well_Test (WT) ─────────────────────────────────
        if "Mapping_Well_Test" in wb.sheetnames:
            ws = wb["Mapping_Well_Test"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                # Same structure as Mapping_Well
                well_name = str(row[3] or "").strip() if len(row) > 3 else ""
                ubhi = str(row[4] or "").strip() if len(row) > 4 else ""
                completion = str(row[5] or "").strip() if len(row) > 5 else ""
                template_type = str(row[6] or "").strip().upper() if len(row) > 6 else "WT"
                attr = str(row[7] or "").strip() if len(row) > 7 else ""
                code = str(row[8] or "").strip() if len(row) > 8 else ""
                ref = str(row[9] or "").strip() if len(row) > 9 and row[9] else ""
                unit = str(row[10] or "").strip() if len(row) > 10 and row[10] else ""

                if code:
                    mappings.append({
                        "template_type": template_type,
                        "well_name": well_name,
                        "ubhi": ubhi,
                        "completion": completion,
                        "attribute_code": code,
                        "attribute": attr,
                        "cell_ref": ref,
                        "unit": unit,
                    })

        # Bulk insert
        if mappings:
            count = config_store.bulk_insert_mappings(concession_id, mappings)
            result.mappings_imported = count

    finally:
        wb.close()

    await logger.success(
        f"Mapping import: {result.mappings_imported} mappings for {concession_id}",
        source="importer",
    )

    return result
