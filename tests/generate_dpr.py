"""Generate DPR Excel files similar to Abir.xlsx with different concession data.
Creates files in docs/dpr/ and configures mappings directly in SQLite.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import random
import sqlite3
import os
import sys

# Paths
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "docs", "dpr")
DB_PATH = os.path.join(os.path.dirname(__file__), "..", "dpr.db")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# 4 fictional concessions with realistic Tunisian oil/gas data
CONCESSIONS = [
    {
        "id": "cherouq_test",
        "name": "CHEROUQ",
        "file_alias": "Cherouq",
        "sheet_name": "CHEROUQ Concession",
        "wells": ["CHR-1", "CHR-2"],
        "oil_range": (120, 350),
        "gas_range": (200000, 500000),
        "water_range": (5, 30),
        "steg_pct": 0.70, "fuel_pct": 0.15, "flare_pct": 0.05,
        "gor_range": (800, 2500), "level": "Ordovicien",
        "choke_range": (20, 40), "fthp_range": (120, 250),
    },
    {
        "id": "sidi_marzoug_test",
        "name": "SIDI MARZOUG",
        "file_alias": "SidiMarzoug",
        "sheet_name": "SIDI MARZOUG Concession",
        "wells": ["SM-3", "SM-5", "SM-7"],
        "oil_range": (50, 180),
        "gas_range": (80000, 250000),
        "water_range": (10, 55),
        "steg_pct": 0.65, "fuel_pct": 0.20, "flare_pct": 0.08,
        "gor_range": (600, 1800), "level": "Silurien Sup.",
        "choke_range": (18, 35), "fthp_range": (80, 190),
    },
    {
        "id": "jebel_grouz_test",
        "name": "JEBEL GROUZ",
        "file_alias": "JebelGrouz",
        "sheet_name": "JEBEL GROUZ Concession",
        "wells": ["JG-1"],
        "oil_range": (200, 600),
        "gas_range": (350000, 800000),
        "water_range": (2, 15),
        "steg_pct": 0.75, "fuel_pct": 0.12, "flare_pct": 0.03,
        "gor_range": (1200, 3500), "level": "Trias",
        "choke_range": (22, 48), "fthp_range": (150, 300),
    },
    {
        "id": "ksar_hadada_test",
        "name": "KSAR HADADA",
        "file_alias": "KsarHadada",
        "sheet_name": "KSAR HADADA Concession",
        "wells": ["KH-2", "KH-4"],
        "oil_range": (80, 250),
        "gas_range": (150000, 400000),
        "water_range": (15, 70),
        "steg_pct": 0.60, "fuel_pct": 0.18, "flare_pct": 0.10,
        "gor_range": (500, 1500), "level": "Permien",
        "choke_range": (16, 32), "fthp_range": (90, 220),
    },
]


def rand(lo, hi):
    return round(random.uniform(lo, hi), 2)


def create_dpr_file(conc):
    """Create a DPR Excel file matching the Abir.xlsx layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = conc["sheet_name"]

    # ── Styles ──
    title_font = Font(name="Arial", size=14, bold=True)
    header_font = Font(name="Arial", size=11, bold=True)
    label_font = Font(name="Arial", size=10)
    value_font = Font(name="Arial", size=10, bold=True)
    name_font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, w in [("B", 35), ("D", 18), ("E", 15), ("F", 20), ("G", 12), ("H", 18)]:
        ws.column_dimensions[col].width = w

    dt = datetime(2024, random.randint(1, 12), random.randint(1, 28))

    # ── Header rows ──
    ws["E2"] = "RAPPORT DE PRODUCTION"
    ws["E2"].font = title_font
    ws["B7"] = "A l'attention de:"
    ws["B7"].font = header_font
    ws["B9"] = "DIRECTION GENERALE DE L'ENERGIE"; ws["H9"] = "XXXXXX"
    ws["B10"] = "ETAP"; ws["H10"] = "XXXXXX"
    ws["B11"] = "OMV"; ws["H11"] = "XXXXXX"
    ws["B12"] = "ATOG"; ws["H12"] = "XXXXXX"

    # Row 15-16 — concession name + date
    ws["B15"] = conc["name"]; ws["B15"].font = name_font
    ws["B16"] = "Production journaliere"; ws["B16"].font = header_font
    ws["F16"] = dt; ws["F16"].number_format = "DD/MM/YYYY"; ws["F16"].font = value_font

    # ── Random production values ──
    oil = rand(*conc["oil_range"])
    gas = rand(*conc["gas_range"])
    water = rand(*conc["water_range"])
    steg = round(gas * conc["steg_pct"])
    fuel = round(gas * conc["fuel_pct"])
    flare = round(gas * conc["flare_pct"])
    pcs = rand(8500, 9800)
    oil_exp = round(oil * rand(0.6, 0.95))
    stock = round(oil * rand(1.5, 4.0))

    # Rows 20-23: Production du champ
    ws["B20"] = "Production du champ:"; ws["B20"].font = header_font
    ws["E20"] = "Jour"; ws["E20"].font = hdr_font; ws["E20"].fill = hdr_fill
    ws["H20"] = "Moy. Mensuelle"; ws["H20"].font = hdr_font; ws["H20"].fill = hdr_fill

    ws["B21"] = "Huile en Baril/Jour"; ws["E21"] = oil; ws["H21"] = round(oil * rand(0.85, 1.15))
    ws["B22"] = "Gaz en SMC/Jour"; ws["E22"] = gas; ws["H22"] = round(gas * rand(0.85, 1.15))
    ws["B23"] = "Eau en Baril/Jour"; ws["E23"] = water; ws["H23"] = round(water * rand(0.85, 1.15), 2)

    # Rows 25-27: Expedition
    ws["B25"] = "Expedition en baril"; ws["B25"].font = header_font
    ws["E25"] = "Jour"; ws["H25"] = "Moy. Mensuelle"
    ws["B26"] = "Huile a Trapsa en bls"; ws["E26"] = oil_exp; ws["H26"] = round(oil_exp * rand(0.9, 1.1))
    ws["B27"] = "Note"; ws["E27"] = "Stock champ (m3) :"; ws["G27"] = stock

    # Rows 29-33: Gaz repartition
    ws["B29"] = "Gaz repartition"; ws["B29"].font = header_font
    ws["E29"] = "Jour"; ws["H29"] = "Moy. Mensuelle"
    ws["B30"] = "Gaz a STEG en SMC"; ws["E30"] = steg; ws["H30"] = round(steg * rand(0.85, 1.15))
    ws["B31"] = "Pouvoir calorifique kcal/Sm3"; ws["E31"] = pcs
    ws["B32"] = "Fuel Gaz Sm3"; ws["E32"] = fuel; ws["H32"] = round(fuel * rand(0.85, 1.15))
    ws["B33"] = "Gaz torche Sm3"; ws["E33"] = flare; ws["H33"] = round(flare * rand(0.85, 1.15))

    # ── Well Parameters (rows 35+) ──
    row = 35
    well_data = {}
    for wn in conc["wells"]:
        ws.cell(row, 2, "Parametres du puits:").font = header_font
        ws.cell(row, 6, wn).font = value_font
        ws.cell(row+1, 2, "Level"); ws.cell(row+1, 6, conc["level"])
        hrs = 24 if random.random() > 0.1 else random.randint(12, 23)
        choke = random.randint(*conc["choke_range"])
        fthp = rand(*conc["fthp_range"])
        w_oil = rand(oil/len(conc["wells"])*0.7, oil/len(conc["wells"])*1.3)
        w_gas = rand(gas/len(conc["wells"])*0.7, gas/len(conc["wells"])*1.3)
        gor = rand(*conc["gor_range"]); bsw = rand(2, 25); dens = rand(0.78, 0.88)

        ws.cell(row+2, 2, "Heure de production"); ws.cell(row+2, 6, hrs)
        ws.cell(row+3, 2, "Choke (fix)"); ws.cell(row+3, 4, "/64 in"); ws.cell(row+3, 6, choke)
        ws.cell(row+4, 2, "FTHP"); ws.cell(row+4, 4, "Min (Bar)"); ws.cell(row+4, 6, fthp)
        ws.cell(row+5, 4, "Max (Bar)"); ws.cell(row+5, 6, round(fthp + rand(1, 5), 2))
        ws.cell(row+6, 2, "Q Oil (net)"); ws.cell(row+6, 4, "(bls/day)"); ws.cell(row+6, 6, w_oil)
        ws.cell(row+7, 2, "Q gaz"); ws.cell(row+7, 4, "(Sm3/day)"); ws.cell(row+7, 6, w_gas)
        ws.cell(row+8, 2, "GOR"); ws.cell(row+8, 4, "(Sm3/m3)"); ws.cell(row+8, 6, gor)
        ws.cell(row+9, 2, "BS & W"); ws.cell(row+9, 4, "(%)"); ws.cell(row+9, 6, bsw)
        ws.cell(row+10, 2, "DENS."); ws.cell(row+10, 4, "(KG/MC)"); ws.cell(row+10, 6, dens)

        well_data[wn] = row
        row += 12

    # ── Well Testing Data ──
    wt_start = row
    for wn in conc["wells"]:
        ws.cell(row, 2, "Well Testing Data").font = header_font
        ws.cell(row, 6, wn).font = value_font
        wt_dt = datetime(2024, random.randint(1, 12), random.randint(1, 28))
        ws.cell(row+1, 2, "Date"); ws.cell(row+1, 6, wt_dt); ws.cell(row+1, 6).number_format = "DD/MM/YYYY"
        ws.cell(row+2, 2, "Wellhead"); ws.cell(row+2, 4, "WHP (kg/cm2)"); ws.cell(row+2, 6, rand(15, 80))
        ws.cell(row+3, 4, "WHT (C)"); ws.cell(row+3, 6, rand(40, 90))
        ws.cell(row+4, 2, "Separator"); ws.cell(row+4, 4, "SP (Kg/cm2)"); ws.cell(row+4, 6, rand(5, 30))
        ws.cell(row+5, 4, "ST(C)"); ws.cell(row+5, 6, rand(30, 60))
        ws.cell(row+6, 2, "Oil"); ws.cell(row+6, 4, "QO (Sm3/d)"); ws.cell(row+6, 6, rand(10, 200))
        ws.cell(row+7, 4, "API"); ws.cell(row+7, 6, rand(28, 45))
        ws.cell(row+8, 4, "Density (g/cm3)"); ws.cell(row+8, 6, rand(0.78, 0.92))
        ws.cell(row+9, 2, "Water"); ws.cell(row+9, 4, "QW (Sm3/d)"); ws.cell(row+9, 6, rand(0.5, 30))
        ws.cell(row+10, 4, "WC (%)"); ws.cell(row+10, 6, rand(2, 35))
        ws.cell(row+11, 2, "Gaz"); ws.cell(row+11, 4, "Qg (Sm3/d)"); ws.cell(row+11, 6, rand(50000, 400000))
        ws.cell(row+12, 4, "GOR (Sm3/d)"); ws.cell(row+12, 6, rand(500, 3000))
        ws.cell(row+13, 4, "SG (air=1)"); ws.cell(row+13, 6, rand(0.6, 1.2))
        row += 15

    # ── Notes ──
    notes_row = row + 1
    ws.cell(row, 2, "NOTES").font = header_font
    ws.cell(notes_row, 2, f"Production {conc['name']} du {dt.strftime('%d/%m/%Y')}. "
                            f"Oil={oil:.0f} bbl, Gas={gas:.0f} Sm3. Operations normales.")

    # Borders
    for r in range(20, notes_row + 1):
        for c in [2, 3, 4, 5, 6, 7, 8]:
            ws.cell(r, c).border = border

    filepath = os.path.join(OUTPUT_DIR, f"{conc['file_alias']}.xlsx")
    wb.save(filepath)
    print(f"  Created: {os.path.basename(filepath)}")
    return well_data, notes_row, wt_start


def configure_db(conc, well_data, notes_row, wt_start):
    """Create concession + mappings in SQLite."""
    conn = sqlite3.connect(DB_PATH)
    cid = conc["id"]

    # Delete if exists
    conn.execute("DELETE FROM cell_mappings WHERE concession_id = ?", (cid,))
    conn.execute("DELETE FROM concessions WHERE id = ?", (cid,))

    # Insert concession
    conn.execute("""
        INSERT INTO concessions (id, name, dpr_file_alias, dpr_sheet,
            active_daily, active_monthly, active_well_test, date_format)
        VALUES (?, ?, ?, ?, 1, 0, 1, 'ddmmyyyy')
    """, (cid, conc["name"], conc["file_alias"], conc["sheet_name"]))

    mappings = []

    # DC mappings
    dc = [
        ("DC001", "Nom Concession", "B:15", ""),
        ("DC002", "Date", "F:16", ""),
        ("DC003", "Remarque", f"B:{notes_row}", ""),
        ("DC005", "Production Gaz en k Sm3", "E:22", "sm3"),
        ("DC007", "Gaz Expedie (Steg) en K Sm3", "E:30", "sm3"),
        ("DC021", "Gaz Torche en k Sm3", "E:33", "sm3"),
        ("DC022", "Fuel Gas Consomme en k Sm3", "E:32", "sm3"),
        ("DC024", "PCS en Kcal/Nm3", "E:31", ""),
        ("DC028", "Production Huile en m3", "E:21", ""),
        ("DC043", "Production Eau en m3", "E:23", ""),
    ]
    for i, (code, attr, ref, unit) in enumerate(dc):
        mappings.append((cid, "DC", "", "", "", code, attr, ref, unit, i))

    # DW mappings (per well)
    dw_i = 0
    for wn, sr in well_data.items():
        dw = [
            ("DW001", "Nom Puits", f"F:{sr}"),
            ("DW002", "Level", f"F:{sr+1}"),
            ("DW003", "Heure de production", f"F:{sr+2}"),
            ("DW004", "Choke (/64 in)", f"F:{sr+3}"),
            ("DW005", "FTHP Min (Bar)", f"F:{sr+4}"),
            ("DW006", "FTHP Max (Bar)", f"F:{sr+5}"),
            ("DW007", "Q Oil net (bls/day)", f"F:{sr+6}"),
            ("DW008", "Q Gaz (Sm3/day)", f"F:{sr+7}"),
            ("DW009", "GOR (Sm3/m3)", f"F:{sr+8}"),
            ("DW010", "BS&W (%)", f"F:{sr+9}"),
            ("DW011", "Densite (KG/MC)", f"F:{sr+10}"),
        ]
        for code, attr, ref in dw:
            mappings.append((cid, "DW", wn, "", "", code, attr, ref, "", dw_i))
            dw_i += 1

    # WT mappings
    wt_i = 0
    wt_row = wt_start
    for wn in conc["wells"]:
        wt = [
            ("WT001", "Well Name", f"F:{wt_row}"),
            ("WT002", "Date", f"F:{wt_row+1}"),
            ("WT003", "WHP (kg/cm2)", f"F:{wt_row+2}"),
            ("WT004", "WHT (C)", f"F:{wt_row+3}"),
            ("WT005", "SP (Kg/cm2)", f"F:{wt_row+4}"),
            ("WT006", "ST (C)", f"F:{wt_row+5}"),
            ("WT007", "QO (Sm3/d)", f"F:{wt_row+6}"),
            ("WT008", "API", f"F:{wt_row+7}"),
            ("WT009", "Density (g/cm3)", f"F:{wt_row+8}"),
            ("WT010", "QW (Sm3/d)", f"F:{wt_row+9}"),
            ("WT011", "WC (%)", f"F:{wt_row+10}"),
            ("WT012", "Qg (Sm3/d)", f"F:{wt_row+11}"),
        ]
        for code, attr, ref in wt:
            mappings.append((cid, "WT", wn, "", "", code, attr, ref, "", wt_i))
            wt_i += 1
        wt_row += 15

    # Bulk insert
    conn.executemany("""
        INSERT INTO cell_mappings
            (concession_id, template_type, well_name, ubhi, completion,
             attribute_code, attribute, cell_ref, unit, sort_order)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, mappings)

    conn.commit()
    conn.close()

    dc_n = sum(1 for m in mappings if m[1] == "DC")
    dw_n = sum(1 for m in mappings if m[1] == "DW")
    wt_n = sum(1 for m in mappings if m[1] == "WT")
    print(f"  Mappings: DC={dc_n}, DW={dw_n}, WT={wt_n} (total={len(mappings)})")


def main():
    print("=== Generating DPR Files & Configuring Mappings ===")
    print(f"Output: {os.path.abspath(OUTPUT_DIR)}")
    print(f"Database: {os.path.abspath(DB_PATH)}\n")

    random.seed(42)

    for conc in CONCESSIONS:
        print(f"\n--- {conc['name']} ({conc['file_alias']}.xlsx) ---")
        well_data, notes_row, wt_start = create_dpr_file(conc)
        configure_db(conc, well_data, notes_row, wt_start)

    print(f"\n\nDONE! Created {len(CONCESSIONS)} DPR files + mappings.")
    print("Files:", ", ".join(f"{c['file_alias']}.xlsx" for c in CONCESSIONS))


if __name__ == "__main__":
    main()
