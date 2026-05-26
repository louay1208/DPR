"""Create a test DPR Excel file + insert matching mappings into dpr.db.

Usage:  uv run python tests/create_test_dpr.py
Result: tests/test_data/TestDPR.xlsx  +  concession 'test_conc' in DB
"""
import sqlite3, sys, json
from pathlib import Path
from datetime import date, datetime

# Ensure project root on path
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Output paths ──────────────────────────────────────────────────────
TEST_DIR = ROOT / "tests" / "test_data"
TEST_DIR.mkdir(parents=True, exist_ok=True)
XLSX_PATH = TEST_DIR / "TestDPR.xlsx"
DB_PATH = ROOT / "dpr.db"
CONC_ID = "test_conc"
CONC_NAME = "TEST CONCESSION"
SHEET_DC = "DC Concession"
SHEET_DW = "DW Wells"
SHEET_MC = "MC Monthly"
SHEET_WT = "WT Well Test"

# ── Styles ────────────────────────────────────────────────────────────
hdr_font = Font(bold=True, size=11)
hdr_fill = PatternFill("solid", fgColor="D9E1F2")
thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

# ── Build workbook ────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ━━━ Sheet 1: DC (Daily Concession) ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_dc = wb.active
ws_dc.title = SHEET_DC

# Row 1-2: Header area
ws_dc["B1"] = "RAPPORT DE PRODUCTION JOURNALIER"
ws_dc["B1"].font = Font(bold=True, size=14)
ws_dc["B2"] = CONC_NAME
ws_dc["B3"] = "Date du DPR:"
ws_dc["E3"] = datetime(2025, 5, 15)
ws_dc["E3"].number_format = "DD/MM/YYYY"

# Row 5: Section header
ws_dc["B5"] = "Production du champ"
ws_dc["E5"] = "Jour"
ws_dc["H5"] = "Cumul Mensuel"
ws_dc["K5"] = "Cumul Annuel"
style_header(ws_dc, 5, 12)

# DC data rows (row 6-25)
dc_layout = [
    (6,  "Production Gaz en k Sm3",          1250.50,   32500.00,  145000.00),
    (7,  "Production Gaz en k Nm3",          1185.25,   30800.00,  137500.00),
    (8,  "Gaz vendu STEG en k Sm3",          800.00,    20800.00,   95000.00),
    (9,  "Gaz vendu STEG en k Nm3",          758.00,    19700.00,   90000.00),
    (10, "Gaz vendu MISKAR en k Sm3",        200.00,     5200.00,   25000.00),
    (11, "Gaz vendu Gabes en k Sm3",          50.00,     1300.00,    6000.00),
    (12, "Pouvoir Calorifique kcal/Sm3",    9850.00,     9850.00,    9850.00),
    (13, "Production Huile Brute en m3",      45.30,     1178.00,    5250.00),
    (14, "Production Eau en m3",              12.50,      325.00,    1450.00),
    (15, "Gaz torché en k Sm3",               15.00,      390.00,    1800.00),
    (16, "Fuel Gaz en k Sm3",                 80.00,     2080.00,    9500.00),
    (17, "Gaz injecté en k Sm3",             105.50,     2743.00,   12500.00),
    (18, "Production GPL en m3",               8.20,      213.00,     980.00),
    (19, "Expédition GPL en m3",               7.80,      202.00,     930.00),
    (20, "Production Butane en Tonnes",        3.10,       80.60,     370.00),
    (21, "Expédition Butane en Tonnes",        2.90,       75.40,     350.00),
    (22, "Production Propane en Tonnes",       2.50,       65.00,     300.00),
    (23, "Expédition Propane en Tonnes",       2.30,       59.80,     280.00),
    (24, "Production Condensat en m3",        18.60,      483.60,    2200.00),
    (25, "Expédition Condensat en m3",        17.90,      465.40,    2100.00),
]

for row_num, label, jour, cum_m, cum_a in dc_layout:
    ws_dc.cell(row_num, 2, label)
    ws_dc.cell(row_num, 5, jour)
    ws_dc.cell(row_num, 8, cum_m)
    ws_dc.cell(row_num, 11, cum_a)

ws_dc.column_dimensions["B"].width = 35
ws_dc.column_dimensions["E"].width = 15
ws_dc.column_dimensions["H"].width = 15
ws_dc.column_dimensions["K"].width = 15

# ━━━ Sheet 2: DW (Daily Well) ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_dw = wb.create_sheet(SHEET_DW)

ws_dw["B1"] = "DONNEES PUITS JOURNALIERES"
ws_dw["B1"].font = Font(bold=True, size=14)
ws_dw["B2"] = CONC_NAME
ws_dw["E2"] = datetime(2025, 5, 15)

# Header row
dw_headers = ["Concession", "Date", "Well Name", "Completion", "Level",
              "Heures Prod", "Gaz Sm3/j", "Huile m3/j", "GOR",
              "Eau m3/j", "BSW %", "Pression Tête Bar"]
for i, h in enumerate(dw_headers, 1):
    ws_dw.cell(4, i, h)
style_header(ws_dw, 4, len(dw_headers))

# Well data rows
wells = [
    (CONC_NAME, datetime(2025,5,15), "WELL-A1",  "A1-T1", "Trias",    24, 450000, 15.2, 2960, 4.1, 21.2, 195),
    (CONC_NAME, datetime(2025,5,15), "WELL-A2",  "A2-T1", "Trias",    22, 380000, 12.8, 2969, 3.5, 21.5, 188),
    (CONC_NAME, datetime(2025,5,15), "WELL-B1",  "B1-O1", "Ordovic",  24, 320000,  8.5, 3765, 2.2, 20.5, 210),
    (CONC_NAME, datetime(2025,5,15), "WELL-B2",  "B2-O1", "Ordovic",  20, 100500,  8.8, 1142, 2.7, 23.5, 175),
]
for idx, w in enumerate(wells):
    r = 5 + idx
    for c, val in enumerate(w, 1):
        ws_dw.cell(r, c, val)

for col_letter in "ABCDEFGHIJKL":
    ws_dw.column_dimensions[col_letter].width = 16

# ━━━ Sheet 3: MC (Monthly Concession) ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_mc = wb.create_sheet(SHEET_MC)

ws_mc["B1"] = "RAPPORT MENSUEL DE PRODUCTION"
ws_mc["B1"].font = Font(bold=True, size=14)

# Monthly data uses columns B-M for Jan-Dec, with labels in col A/B
# Row 3: headers
ws_mc["B3"] = CONC_NAME
ws_mc["B4"] = "Mois:"
# Columns: C=Jan, D=Feb, ... N=Dec  (for month-swapping logic)
months = ["Jan","Fev","Mar","Avr","Mai","Jun","Jul","Aou","Sep","Oct","Nov","Dec"]
for i, m in enumerate(months):
    ws_mc.cell(4, 3 + i, m)
style_header(ws_mc, 4, 14)

mc_layout = [
    (6,  "Production Gaz k Sm3",      [31000,29500,33000,31500,32500,30000,34000,33500,31000,32000,30500,33000]),
    (7,  "Production Gaz k Nm3",      [29300,27900,31200,29800,30800,28400,32200,31700,29300,30300,28900,31200]),
    (8,  "Gaz vendu STEG k Sm3",      [20000,19000,21300,20300,20800,19400,21900,21600,20000,20600,19700,21300]),
    (9,  "Production Huile m3",        [1100, 1050, 1200, 1140, 1178, 1080, 1230, 1210, 1100, 1150, 1090, 1200]),
    (10, "Production Eau m3",          [310,  295,  340,  322,  325,  300,  345,  340,  310,  325,  305,  340]),
    (11, "Gaz torché k Sm3",           [370,  355,  400,  380,  390,  360,  410,  405,  370,  385,  365,  400]),
    (12, "Production Condensat m3",    [460,  440,  500,  475,  484,  450,  510,  502,  460,  480,  455,  500]),
]

for row_num, label, values in mc_layout:
    ws_mc.cell(row_num, 2, label)
    for i, v in enumerate(values):
        ws_mc.cell(row_num, 3 + i, v)

ws_mc.column_dimensions["B"].width = 30

# ━━━ Sheet 4: WT (Well Test) ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_wt = wb.create_sheet(SHEET_WT)

ws_wt["B1"] = "DONNEES WELL TEST"
ws_wt["B1"].font = Font(bold=True, size=14)

# Well test data laid out vertically per well (columns F, I for 2 wells)
wt_labels = [
    (3, "Concession"),
    (4, "Date du Test"),
    (5, "Nom du Puits"),
    (6, "Completion"),
    (7, "Level"),
    (8, "Choke (/64 in)"),
    (9, "FTHP Min (Bar)"),
    (10, "FTHP Max (Bar)"),
    (11, "Q Oil net (m3/j)"),
    (12, "Q Gaz (Sm3/j)"),
    (13, "GOR (Sm3/m3)"),
    (14, "BS&W (%)"),
    (15, "Densité (g/cm3)"),
    (16, "Q Eau (m3/j)"),
    (17, "WHP (kg/cm2)"),
    (18, "WHT (°C)"),
]

for row_num, label in wt_labels:
    ws_wt.cell(row_num, 2, label)

# Well A1 test data (column F)
ws_wt.cell(3, 6, CONC_NAME)
ws_wt.cell(4, 6, datetime(2025, 4, 20))
ws_wt.cell(5, 6, "WELL-A1")
ws_wt.cell(6, 6, "A1-T1")
ws_wt.cell(7, 6, "Trias")
ws_wt.cell(8, 6, 28)
ws_wt.cell(9, 6, 192)
ws_wt.cell(10, 6, 198)
ws_wt.cell(11, 6, 15.5)
ws_wt.cell(12, 6, 455000)
ws_wt.cell(13, 6, 2935)
ws_wt.cell(14, 6, 21.0)
ws_wt.cell(15, 6, 0.832)
ws_wt.cell(16, 6, 4.2)
ws_wt.cell(17, 6, 195)
ws_wt.cell(18, 6, 82)

# Well B1 test data (column I)
ws_wt.cell(3, 9, CONC_NAME)
ws_wt.cell(4, 9, datetime(2025, 4, 22))
ws_wt.cell(5, 9, "WELL-B1")
ws_wt.cell(6, 9, "B1-O1")
ws_wt.cell(7, 9, "Ordovic")
ws_wt.cell(8, 9, 22)
ws_wt.cell(9, 9, 205)
ws_wt.cell(10, 9, 215)
ws_wt.cell(11, 9, 8.8)
ws_wt.cell(12, 9, 325000)
ws_wt.cell(13, 9, 3693)
ws_wt.cell(14, 9, 20.0)
ws_wt.cell(15, 9, 0.845)
ws_wt.cell(16, 9, 2.2)
ws_wt.cell(17, 9, 212)
ws_wt.cell(18, 9, 78)

ws_wt.column_dimensions["B"].width = 25
ws_wt.column_dimensions["F"].width = 18
ws_wt.column_dimensions["I"].width = 18

# ── Save workbook ─────────────────────────────────────────────────────
wb.save(XLSX_PATH)
print(f"✓ Created {XLSX_PATH}")

# ══════════════════════════════════════════════════════════════════════
# INSERT MAPPINGS INTO DB
# ══════════════════════════════════════════════════════════════════════
conn = sqlite3.connect(str(DB_PATH))
conn.execute("PRAGMA foreign_keys=ON")

# Delete old test concession if exists
conn.execute("DELETE FROM cell_mappings WHERE concession_id=?", (CONC_ID,))
conn.execute("DELETE FROM concessions WHERE id=?", (CONC_ID,))

# Insert concession
conn.execute("""
    INSERT INTO concessions (id, name, dpr_file_alias, dpr_sheet,
        active_daily, active_monthly, active_well_test, monthly_report)
    VALUES (?, ?, ?, ?, 1, 1, 1, '')
""", (CONC_ID, CONC_NAME, "TestDPR", SHEET_DC))

# ── DC mappings (cell_ref = "COL:ROW") ────────────────────────────
dc_mappings = [
    ("DC001", "Nom Concession",              "B2",  ""),
    ("DC002", "Date du DPR",                 "E3",  ""),
    ("DC005", "Production Gaz en k Sm3",     "E6",  ""),
    ("DC006", "Production Gaz en k Nm3",     "E7",  ""),
    ("DC007", "Gaz vendu STEG en k Sm3",     "E8",  ""),
    ("DC009", "Gaz vendu MISKAR en k Sm3",   "E10", ""),
    ("DC011", "Gaz vendu Gabes en k Sm3",    "E11", ""),
    ("DC021", "Gaz torché en k Sm3",         "E15", ""),
    ("DC022", "Fuel Gaz en k Sm3",           "E16", ""),
    ("DC023", "Gaz injecté en k Sm3",        "E17", ""),
    ("DC028", "Production Huile Brute m3",   "E13", ""),
    ("DC031", "Production GPL en m3",        "E18", ""),
    ("DC032", "Expédition GPL en m3",        "E19", ""),
    ("DC034", "Production Butane Tonnes",    "E20", ""),
    ("DC035", "Expédition Butane Tonnes",    "E21", ""),
    ("DC037", "Production Propane Tonnes",   "E22", ""),
    ("DC038", "Expédition Propane Tonnes",   "E23", ""),
    ("DC043", "Production Eau en m3",        "E14", ""),
    ("DC047", "Production Condensat m3",     "E24", ""),
    ("DC048", "Expédition Condensat m3",     "E25", ""),
]

# ── DW mappings (well_name + cell_ref per well) ──────────────────
# Wells are in rows 5-8, columns A-L
dw_base = [
    ("DW001", "Concession",     1, ""),
    ("DW002", "Date",           2, ""),
    ("DW003", "Nom du Puits",   3, ""),
    ("DW004", "Completion",     4, ""),
    ("DW005", "Level",          5, ""),
    ("DW006", "Heures Prod",    6, ""),
    ("DW007", "Gaz Sm3/j",      7, ""),
    ("DW008", "Huile m3/j",     8, ""),
    ("DW009", "GOR",            9, ""),
    ("DW010", "Eau m3/j",       10, ""),
    ("DW011", "BSW %",          11, ""),
    ("DW012", "Pression Tête",  12, ""),
]

dw_wells = [
    ("WELL-A1", "A1-T1", 5),
    ("WELL-A2", "A2-T1", 6),
    ("WELL-B1", "B1-O1", 7),
    ("WELL-B2", "B2-O1", 8),
]

# ── MC mappings ───────────────────────────────────────────────────
mc_mappings = [
    ("MC001", "Nom Concession",           "B3",  ""),
    ("MC002", "Date",                     "",    ""),
    ("MC005", "Production Gaz k Sm3",     "C6",  ""),
    ("MC006", "Production Gaz k Nm3",     "C7",  ""),
    ("MC007", "Gaz vendu STEG k Sm3",     "C8",  ""),
    ("MC028", "Production Huile m3",      "C9",  ""),
    ("MC043", "Production Eau m3",        "C10", ""),
    ("MC021", "Gaz torché k Sm3",         "C11", ""),
    ("MC047", "Production Condensat m3",  "C12", ""),
]

# ── WT mappings (per well) ────────────────────────────────────────
wt_base = [
    ("WT001", "Concession",     3, ""),
    ("WT002", "Date du Test",   4, ""),
    ("WT003", "Nom du Puits",   5, ""),
    ("WT004", "Completion",     6, ""),
    ("WT005", "Level",          7, ""),
    ("WT006", "Choke",          8, ""),
    ("WT007", "FTHP Min",       9, ""),
    ("WT008", "FTHP Max",      10, ""),
    ("WT009", "Q Oil net",     11, ""),
    ("WT010", "Q Gaz",         12, ""),
    ("WT011", "GOR",           13, ""),
    ("WT012", "BSW",           14, ""),
    ("WT013", "Densité",       15, ""),
    ("WT014", "Q Eau",         16, ""),
]

wt_wells = [
    ("WELL-A1", "A1-T1", "F"),  # column F
    ("WELL-B1", "B1-O1", "I"),  # column I
]

# ── Insert all mappings ───────────────────────────────────────────
insert_sql = """
    INSERT INTO cell_mappings
        (concession_id, template_type, well_name, ubhi, completion,
         attribute_code, attribute, cell_ref, unit, sort_order)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
"""

order = 0

# DC
for code, attr, ref, unit in dc_mappings:
    order += 1
    conn.execute(insert_sql, (CONC_ID, "DC", "", "", "", code, attr, ref, unit, order))

# DW
for well_name, completion, row_num in dw_wells:
    for code, attr, col_idx, unit in dw_base:
        order += 1
        col_letter = chr(64 + col_idx)  # 1=A, 2=B, ...
        ref = f"{col_letter}{row_num}"
        conn.execute(insert_sql, (CONC_ID, "DW", well_name, "", completion,
                                  code, attr, ref, unit, order))

# MC
for code, attr, ref, unit in mc_mappings:
    order += 1
    conn.execute(insert_sql, (CONC_ID, "MC", "", "", "", code, attr, ref, unit, order))

# WT
for well_name, completion, col in wt_wells:
    for code, attr, row_num, unit in wt_base:
        order += 1
        ref = f"{col}{row_num}"
        conn.execute(insert_sql, (CONC_ID, "WT", well_name, "", completion,
                                  code, attr, ref, unit, order))

# ── Insert naming rule so the parser can find 'TestDPR' ───────────
conn.execute("DELETE FROM naming_rules WHERE alias='TestDPR'")
conn.execute("""
    INSERT INTO naming_rules (alias, file_alias, extension, date_format)
    VALUES ('TestDPR', 'TestDPR', 'xlsx', 'ddmmyyyy')
""")

# ── Set dpr_folder parameter ──────────────────────────────────────
conn.execute("INSERT OR REPLACE INTO parameters (key, value) VALUES ('dpr_folder', ?)",
             (str(TEST_DIR),))

conn.commit()

# Verify
total = conn.execute("SELECT COUNT(*) FROM cell_mappings WHERE concession_id=?", (CONC_ID,)).fetchone()[0]
print(f"✓ Inserted concession '{CONC_NAME}' with {total} mappings")
print(f"  DC: {len(dc_mappings)}, DW: {len(dw_base) * len(dw_wells)}, MC: {len(mc_mappings)}, WT: {len(wt_base) * len(wt_wells)}")
print(f"✓ dpr_folder set to: {TEST_DIR}")
print(f"\n→ Restart the server, then extract with report_type=daily to test DC+DW")
print(f"→ Use report_type=monthly for MC, report_type=well_test for WT")

conn.close()
